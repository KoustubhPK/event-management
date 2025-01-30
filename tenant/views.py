from django.shortcuts import render, redirect, get_object_or_404
from .models import Guest, Event, Gift, OfferedGift, ReceivedGift
from .forms import GuestForm, EventForm, GiftForm, OfferedGiftForm, ReceivedGiftForm
from django.db import transaction
from django.db.models import Count, Prefetch, Subquery, F, Q
from django.http import HttpResponse
from django.contrib import messages

import openpyxl
from openpyxl.utils import get_column_letter

# Create your views here.

def index(request):
    total_guests = Guest.objects.count()
    total_events = Event.objects.count()
    unattended_guests = Guest.objects.filter(event__isnull=True).count()
    
    # Calculate the percentile of guests who have attended at least one event
    attended_guests_count = total_guests - unattended_guests
    attended_guests_percentile = (attended_guests_count / max(total_guests, 1)) * 100
    # Query for guests who have attended at least one event
    attended_guests_query = Guest.objects.annotate(num_events=Count('event', distinct=True)).filter(num_events__gt=0)
    # Fetch attended guests
    attended_guests = attended_guests_query.all()

    # Calculate the count and percentage of least involved guest
    least_involved_guest_query = Guest.objects.annotate(num_events=Count('event', distinct=True))
    least_involved_guest_subquery = least_involved_guest_query.order_by('num_events').values('num_events')[:1]
    least_involved_guest_count = least_involved_guest_query.filter(num_events=Subquery(least_involved_guest_subquery)).count()
    least_involved_guest_percentage = (least_involved_guest_count / max(total_guests, 1)) * 100
    least_involved_guests = least_involved_guest_query.filter(num_events=Subquery(least_involved_guest_subquery))

    # Calculate the count and percentage of most involved guest
    most_involved_guest_query = Guest.objects.annotate(num_events=Count('event', distinct=True))
    most_involved_guest_subquery = most_involved_guest_query.order_by('-num_events').values('num_events')[:1]
    most_involved_guest_count = most_involved_guest_query.filter(num_events=Subquery(most_involved_guest_subquery)).count()
    most_involved_guest_percentage = (most_involved_guest_count / max(total_guests, 1)) * 100
    most_involved_guests = most_involved_guest_query.filter(num_events=Subquery(most_involved_guest_subquery))
    
    # Data for the bar chart
    events_with_guest_count = Event.objects.annotate(guest_count=Count('guests'))
    event_names = [event.name for event in events_with_guest_count]
    guest_counts = [event.guest_count for event in events_with_guest_count]

    context = {
        'total_guests': total_guests,
        'attended_guests_count': attended_guests_count,
        'attended_guests_percentile': attended_guests_percentile,
        'attended_guests': attended_guests,
        'total_events': total_events,
        'unattended_guests': unattended_guests,
        'least_involved_guests': least_involved_guests,
        'least_involved_guest_count': least_involved_guest_count,
        'least_involved_guest_percentage': least_involved_guest_percentage,
        'most_involved_guests': most_involved_guests,
        'most_involved_guest_count': most_involved_guest_count,
        'most_involved_guest_percentage': most_involved_guest_percentage,
        'event_names': event_names,
        'guest_counts': guest_counts,
    }
    return render(request, 'index.html', context)

def guest_list(request):
    events = Event.objects.all()
    guests = Guest.objects.prefetch_related(Prefetch('event_set', queryset=events, to_attr='events'))
    # guests = Guest.objects.annotate(event_count=Count('event'))
    context = {'guests': guests}
    return render(request, 'guests/guest_list.html', context)

def export_guests_xlsx(request):
    guests = Guest.objects.annotate(event_count=Count('event'))

    # Create a workbook and add a worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Guests'

    # Define the headers
    headers = ['No.', 'Name', 'Events']
    ws.append(headers)

    # Add the guest data
    for idx, guest in enumerate(guests, start=1):
        ws.append([idx, guest.name, guest.event_count])

    # Adjust column widths
    for col_num, col in enumerate(ws.columns, start=1):
        max_length = 0
        column = get_column_letter(col_num)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Create a response object
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=guests.xlsx'

    # Save the workbook to the response
    wb.save(response)

    return response

def add_guest(request):
    if request.method == "POST":
        form = GuestForm(request.POST)
        if form.is_valid():
            name = form.cleaned_data['name']
            if Guest.objects.filter(name=name).exists():
                messages.warning(request, 'A guest with this name already exists.')
            else:
                form.save()
                messages.success(request, 'Guest added successfully.')
                return redirect('add_guest')
    else:
        form = GuestForm()
    return render(request, 'guests/add_guest.html', {'form': form})

def edit_guest(request, pk):
    guest = get_object_or_404(Guest, pk=pk)
    if request.method == "POST":
        form = GuestForm(request.POST, instance=guest)
        if form.is_valid():
            form.save()
            return redirect('guest_list')
    else:
        form = GuestForm(instance=guest)
    return render(request, 'guests/edit_guest.html', {'form': form})

def delete_guest(request, pk):
    guest = get_object_or_404(Guest, pk=pk)
    if request.method == "POST":
        guest.delete()
        return redirect('guest_list')
    return render(request, 'guests/delete_guest.html', {'guest': guest})


def add_event(request):
    if request.method == "POST":
        form = EventForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('event_list')  # Adjust this to your event list view name
    else:
        form = EventForm()
    return render(request, 'events/add_event.html', {'form': form})

def event_list(request):
    events = Event.objects.all()
    return render(request, 'events/event_list.html', {'events': events})

def event_detail(request, pk):
    event = get_object_or_404(Event, pk=pk)
    guest_count = event.guests.count()
    return render(request, 'events/event_detail.html', {'event': event, 'guest_count': guest_count})

def edit_event(request, pk):
    event = get_object_or_404(Event, pk=pk)
    if request.method == "POST":
        form = EventForm(request.POST, instance=event)
        if form.is_valid():
            form.save()
            return redirect('event_detail', pk=event.pk)
    else:
        form = EventForm(instance=event)
    return render(request, 'events/edit_event.html', {'form': form})

def remove_guest_from_event(request, event_pk, guest_pk):
    event = get_object_or_404(Event, pk=event_pk)
    guest = get_object_or_404(Guest, pk=guest_pk)
    event.guests.remove(guest)
    return redirect('event_detail', pk=event_pk)

def delete_event(request, pk):
    event = get_object_or_404(Event, pk=pk)
    if request.method == "POST":
        event.delete()
        return redirect('event_list')
    return render(request, 'events/delete_event.html', {'event': event})

def add_gift(request):
    if request.method == "POST":
        form = GiftForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, "Gift added successfully.")
            return redirect('gift_list')
    else:
        form = GiftForm()
    return render(request, 'gifts/add_gift.html', {'form': form})

def edit_gift(request, pk):
    gift = get_object_or_404(Gift, pk=pk)
    if request.method == "POST":
        form = GiftForm(request.POST, request.FILES, instance=gift)
        if form.is_valid():
            form.save()
            messages.success(request, "Gift updated successfully.")
            return redirect('gift_list')
    else:
        form = GiftForm(instance=gift)
    return render(request, 'gifts/edit_gift.html', {'form': form})

def delete_gift(request, pk):
    gift = get_object_or_404(Gift, pk=pk)
    if request.method == "POST":
        gift.delete()
        messages.success(request, "Gift deleted successfully.")
        return redirect('gift_list')
    return render(request, 'gifts/delete_gift.html', {'gift': gift})

def gift_list(request):
    gifts = Gift.objects.all()
    offered_gifts = OfferedGift.objects.all()
    received_gifts = ReceivedGift.objects.all()

    # Data for the bar chart
    gifts_with_guest_count = OfferedGift.objects.annotate(guest_count=Count('guests'))
    gift_names = [gift.gift.name for gift in gifts_with_guest_count]
    guest_counts = [gift.guest_count for gift in gifts_with_guest_count]

    # Calculate the total value for each gift
    for gift in gifts:
        offered_count = offered_gifts.filter(gift=gift).aggregate(total_offered=Count('guests'))['total_offered']
        gift.total = gift.stock
        gift.available_stock = gift.stock - offered_count if offered_count else gift.stock
    
    context = {
        'gifts': gifts,
        'offered_gifts': offered_gifts,
        'received_gifts': received_gifts,

        'gift_names': gift_names,
        'guest_counts': guest_counts,
    }

    return render(request, 'gifts/gift_list.html', context)

def offered_gift_detail(request, pk):
    gift = get_object_or_404(OfferedGift, pk=pk)
    guest_count = gift.guests.count()
    return render(request, 'gifts/offered_gift_detail.html', {'gift': gift, 'guest_count': guest_count})

def remove_guest_from_offered_gift(request, gift_pk, guest_pk):
    offered_gift = get_object_or_404(OfferedGift, pk=gift_pk)
    guest = get_object_or_404(Guest, pk=guest_pk)
    offered_gift.guests.remove(guest)
    messages.success(request, f"Guest {guest.name} removed from the gift {offered_gift.gift.name}.")
    return redirect('offered_gift_detail', pk=gift_pk)

# @transaction.atomic
# def add_offer_gift(request):
#     if request.method == "POST":
#         form = OfferedGiftForm(request.POST)
#         if form.is_valid():
#             gift = form.cleaned_data['gift']
#             guests = form.cleaned_data['guests']

#             # Calculate the number of guests already offered this gift
#             already_offered_count = OfferedGift.objects.filter(gift=gift).aggregate(total_offered=Count('guests'))['total_offered']
#             available_stock = gift.stock - already_offered_count

#             if available_stock < len(guests):
#                 messages.error(request, "Not enough stock available for the selected gift.")
#             else:
#                 offered_gift = form.save()
#                 messages.success(request, "Gift offered successfully.")
#                 return redirect('gift_list')
#     else:
#         form = OfferedGiftForm()

#     return render(request, 'gifts/add_offer_gift.html', {'form': form})

@transaction.atomic
def add_offer_gift(request):
    # Get gifts that have not been offered to any guests
    gifts_with_no_offers = Gift.objects.annotate(
        offered_count=Count('offered_gifts__guests')
    ).filter(
        offered_count=0
    )

    if not gifts_with_no_offers.exists():
        return render(request, 'gifts/add_offer_gift.html', {'no_stock': True})

    if request.method == "POST":
        form = OfferedGiftForm(request.POST)
        if form.is_valid():
            offered_gift = form.save(commit=False)
            if offered_gift.gift.offered_gifts.exists():
                messages.error(request, "This gift has already been offered.")
            else:
                offered_gift.save()
                form.save_m2m()  # Save the many-to-many relationships
                messages.success(request, "Gift offered successfully.")
                return redirect('gift_list')
    else:
        form = OfferedGiftForm()

    # Limit the gift choices in the form to only those with no offers
    form.fields['gift'].queryset = gifts_with_no_offers

    return render(request, 'gifts/add_offer_gift.html', {'form': form})

def edit_offered_gift(request, pk):
    offered_gift = get_object_or_404(OfferedGift, pk=pk)
    
    if request.method == "POST":
        form = OfferedGiftForm(request.POST, instance=offered_gift)
        if form.is_valid():
            form.save()
            messages.success(request, "Offered Gift updated successfully.")
            return redirect('gift_list')
    else:
        form = OfferedGiftForm(instance=offered_gift)
    
    return render(request, 'gifts/edit_offered_gift.html', {'form': form})

def delete_offered_gift(request, pk):
    offered_gift = get_object_or_404(OfferedGift, pk=pk)
    if request.method == "POST":
        offered_gift.delete()
        messages.success(request, "Offered Gift deleted successfully.")
        return redirect('gift_list')
    
    return render(request, 'gifts/delete_offered_gift.html', {'offered_gift': offered_gift})

def add_receive_gift(request):
    if request.method == "POST":
        form = ReceivedGiftForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, "Gift received successfully.")
            return redirect('gift_list')
    else:
        form = ReceivedGiftForm()
    return render(request, 'gifts/add_receive_gift.html', {'form': form})
